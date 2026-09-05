from __future__ import annotations

from pathlib import Path
import re
from typing import Any

import pandas as pd

from .dates import parse_datetime_value


HOLIDAY_ALIASES = {
    "state": ["state", "territory", "region", "work state", "jurisdiction"],
    "holiday_date": ["holiday date", "public holiday date", "date of holiday", "date"],
    "holiday_name": ["holiday name", "public holiday name", "holiday", "description", "name"],
    "holiday_location_key": ["holiday location key", "location key", "local holiday key", "location", "town", "region key", "site"],
    "holiday_scope": ["holiday scope", "scope"],
    "source": ["source", "reference", "evidence source"],
}

STATE_ALIASES = {
    "WA": "WA", "WESTERN AUSTRALIA": "WA",
    "NSW": "NSW", "NEW SOUTH WALES": "NSW",
    "VIC": "VIC", "VICTORIA": "VIC",
    "QLD": "QLD", "QUEENSLAND": "QLD",
    "SA": "SA", "SOUTH AUSTRALIA": "SA",
    "TAS": "TAS", "TASMANIA": "TAS",
    "ACT": "ACT", "AUSTRALIAN CAPITAL TERRITORY": "ACT",
    "NT": "NT", "NORTHERN TERRITORY": "NT",
}

VALID_SCOPES = {"STATEWIDE", "LOCAL"}


def _norm(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", " ", str(value or "").lower()).strip()


def _blank(value: Any) -> bool:
    if value is None:
        return True
    try:
        if pd.isna(value):
            return True
    except Exception:
        pass
    return not str(value).strip()


def _find_column(columns: list[str], aliases: list[str]) -> str | None:
    normalized = {_norm(column): column for column in columns}
    for alias in aliases:
        key = _norm(alias)
        if key in normalized:
            return normalized[key]
    for column in columns:
        c = _norm(column)
        for alias in aliases:
            a = _norm(alias)
            if a and set(a.split()).issubset(set(c.split())):
                return column
    return None


def _read_items(root: Path):
    for path in sorted(root.rglob("*")):
        if not path.is_file() or path.name.startswith("~$"):
            continue
        if path.name.lower().startswith("public_holiday_overrides_template"):
            continue
        rel = str(path.relative_to(root)).replace("\\", "/")
        if path.suffix.lower() == ".csv":
            yield rel, None, pd.read_csv(path)
        elif path.suffix.lower() in {".xlsx", ".xlsm"}:
            book = pd.ExcelFile(path)
            for sheet in book.sheet_names:
                yield rel, sheet, pd.read_excel(path, sheet_name=sheet)


def _is_holiday_item(file: str, sheet: str | None, columns: list[str], mapped: dict[str, str | None]) -> bool:
    if not (mapped.get("state") and mapped.get("holiday_date") and mapped.get("holiday_name")):
        return False
    label = _norm(f"{file} {sheet or ''}")
    explicit_holiday_column = any("holiday" in _norm(column) for column in columns)
    return "holiday" in label or explicit_holiday_column


def _state(value: Any) -> str | None:
    if _blank(value):
        return None
    return STATE_ALIASES.get(str(value).strip().upper())


def detect_public_holiday_overrides(source_root: str | Path) -> dict[str, Any]:
    """Detect and canonicalise operator-supplied one-off/local public holiday files.

    Detection is conservative: a source must contain state, holiday date and holiday
    name fields and either identify itself as a holiday source by file/sheet name or
    contain an explicitly holiday-named column.
    """
    root = Path(source_root)
    if not root.exists():
        return {"frame": pd.DataFrame(), "interpretation": [], "warnings": []}

    rows: list[dict[str, Any]] = []
    interpretation: list[dict[str, Any]] = []
    warnings: list[str] = []

    for file, sheet, frame in _read_items(root):
        columns = list(frame.columns)
        mapped = {field: _find_column(columns, aliases) for field, aliases in HOLIDAY_ALIASES.items()}
        if not _is_holiday_item(file, sheet, columns, mapped):
            continue
        label = file if sheet is None else f"{file}#{sheet}"
        interpretation.append({
            "source_file": file,
            "source_sheet": sheet,
            "detected_roles": ["public_holiday_overrides"],
            "mapped_fields": {k: v for k, v in mapped.items() if v},
            "sample_rows": len(frame),
        })
        for row_no, row in frame.iterrows():
            raw_date = row.get(mapped["holiday_date"])
            raw_name = row.get(mapped["holiday_name"])
            raw_state = row.get(mapped["state"])
            if _blank(raw_date) and _blank(raw_name) and _blank(raw_state):
                continue
            date = parse_datetime_value(raw_date)
            if pd.isna(date):
                raise ValueError(f"{label} row {int(row_no)+2}: invalid holiday date {raw_date!r}. Use DD-MM-YYYY, DD/MM/YYYY or ISO YYYY-MM-DD.")
            state = _state(raw_state)
            if state is None:
                raise ValueError(f"{label} row {int(row_no)+2}: unsupported or missing Australian state/territory {raw_state!r}.")
            if _blank(raw_name):
                raise ValueError(f"{label} row {int(row_no)+2}: holiday name is required.")
            location = row.get(mapped["holiday_location_key"]) if mapped.get("holiday_location_key") else None
            location = None if _blank(location) else str(location).strip()
            scope_raw = row.get(mapped["holiday_scope"]) if mapped.get("holiday_scope") else None
            scope = str(scope_raw).strip().upper() if not _blank(scope_raw) else ("LOCAL" if location else "STATEWIDE")
            if scope not in VALID_SCOPES:
                raise ValueError(f"{label} row {int(row_no)+2}: holiday scope must be STATEWIDE or LOCAL, got {scope_raw!r}.")
            if scope == "LOCAL" and not location:
                raise ValueError(f"{label} row {int(row_no)+2}: LOCAL holiday requires a holiday_location_key/location value.")
            source_raw = row.get(mapped["source"]) if mapped.get("source") else None
            source = "manual_override" if _blank(source_raw) else str(source_raw).strip()
            rows.append({
                "state": state,
                "holiday_date": date.normalize(),
                "holiday_name": str(raw_name).strip(),
                "holiday_location_key": location,
                "holiday_scope": scope,
                "source": source,
                "source_file": file,
                "source_sheet": sheet,
                "source_row": int(row_no) + 2,
            })

    result = pd.DataFrame(rows)
    if not result.empty:
        duplicate_mask = result.duplicated(subset=["state", "holiday_date", "holiday_name", "holiday_location_key"], keep="first")
        if duplicate_mask.any():
            warnings.append(f"Ignored {int(duplicate_mask.sum())} duplicate public holiday override row(s).")
            result = result.loc[~duplicate_mask].copy()
        warnings.append(f"Detected {len(result)} operator-supplied public holiday override row(s).")
    return {"frame": result, "interpretation": interpretation, "warnings": warnings}


def write_public_holiday_template(path: str | Path) -> Path:
    """Create an operator Excel template with controlled dropdowns."""
    from openpyxl import Workbook
    from openpyxl.worksheet.datavalidation import DataValidation

    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "public_holiday_overrides"
    headers = ["state", "holiday_date", "holiday_name", "holiday_location_key", "holiday_scope", "source"]
    ws.append(headers)
    ws.append(["WA", "15-08-2025", "Example Special Public Holiday", None, "STATEWIDE", "manual_override"])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = "A1:F500"
    widths = {"A": 16, "B": 18, "C": 42, "D": 28, "E": 18, "F": 28}
    for column, width in widths.items():
        ws.column_dimensions[column].width = width

    lists = wb.create_sheet("_lists")
    lists.sheet_state = "hidden"
    states = ["WA", "NSW", "VIC", "QLD", "SA", "TAS", "ACT", "NT"]
    scopes = ["STATEWIDE", "LOCAL"]
    sources = ["manual_override", "government_notice", "gazette", "employer_calendar", "other_evidence"]
    for idx, value in enumerate(states, 1):
        lists.cell(idx, 1, value)
    for idx, value in enumerate(scopes, 1):
        lists.cell(idx, 2, value)
    for idx, value in enumerate(sources, 1):
        lists.cell(idx, 3, value)

    state_dv = DataValidation(type="list", formula1="'_lists'!$A$1:$A$8", allow_blank=False)
    scope_dv = DataValidation(type="list", formula1="'_lists'!$B$1:$B$2", allow_blank=False)
    source_dv = DataValidation(type="list", formula1="'_lists'!$C$1:$C$5", allow_blank=True)
    for dv, target in ((state_dv, "A2:A500"), (scope_dv, "E2:E500"), (source_dv, "F2:F500")):
        dv.error = "Select a value from the dropdown list."
        dv.errorTitle = "Invalid AuditHero value"
        dv.showErrorMessage = True
        ws.add_data_validation(dv)
        dv.add(target)

    readme = wb.create_sheet("README", 0)
    instructions = [
        ["AuditHero Public Holiday Override Template"],
        ["Purpose", "Use only for one-off, irregular, local/regional or otherwise missing public holidays. Normal Australian statewide public holidays are generated automatically."],
        ["Date format", "Enter holiday_date as DD-MM-YYYY (for example 15-08-2025). ISO YYYY-MM-DD is also accepted."],
        ["STATEWIDE", "Leave holiday_location_key blank."],
        ["LOCAL", "Enter a stable holiday_location_key such as KUNUNURRA. The same key must be present on affected timesheet rows."],
        ["State", "Choose WA, NSW, VIC, QLD, SA, TAS, ACT or NT from the dropdown."],
        ["Source", "Choose the evidence source where practical. You may also use your own source text in ordinary CSV/XLSX uploads; AuditHero preserves it as evidence context."],
        ["Normal workflow", "Upload this workbook with the other raw source files, run AuditHero - Preview Uploaded Files, confirm it is detected as public_holiday_overrides, then run AuditHero - Audit Reviewed Uploaded Files."],
    ]
    for row in instructions:
        readme.append(row)
    readme.column_dimensions["A"].width = 24
    readme.column_dimensions["B"].width = 120
    wb.save(path)
    return path
