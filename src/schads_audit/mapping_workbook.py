from __future__ import annotations

from pathlib import Path
from typing import Any
import json

import pandas as pd

from .source_mapping import CANONICAL_SCHEMAS


SPECIAL_FILE_ROLES = ["CONTEXT_ONLY", "IGNORE"]


def _bool(value: Any, default: bool = False) -> bool:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return default
    if isinstance(value, bool):
        return value
    return str(value).strip().lower() in {"true", "1", "yes", "y", "on"}


def _text(value: Any) -> str | None:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None
    text = str(value).strip()
    return text if text else None


def _split(value: Any) -> list[str]:
    text = _text(value)
    if not text:
        return []
    return [x.strip() for x in text.split(";") if x.strip()]


def _inventory(draft: dict[str, Any]) -> list[dict[str, Any]]:
    rows = draft.get("_source_inventory") or []
    cleaned: list[dict[str, Any]] = []
    for row in rows:
        if not isinstance(row, dict):
            continue
        cleaned.append({
            "file": _text(row.get("file")),
            "sheet": _text(row.get("sheet")),
            "item_name": _text(row.get("item_name")),
            "sample_rows": row.get("sample_rows"),
            "columns": list(row.get("columns") or []),
        })
    return cleaned


def _detected_role_for_item(draft: dict[str, Any], file: str | None, sheet: str | None) -> tuple[str | None, float | None]:
    best_role, best_conf = None, -1.0
    for dataset, cfg in (draft.get("datasets") or {}).items():
        source = (cfg or {}).get("source") or {}
        if _text(source.get("file")) != file or _text(source.get("sheet")) != sheet:
            continue
        confidence = float((cfg or {}).get("_dataset_confidence") or 0.0)
        if confidence > best_conf:
            best_role, best_conf = dataset, confidence
    return best_role, (round(best_conf, 3) if best_role else None)


def draft_to_field_table(draft: dict[str, Any]) -> pd.DataFrame:
    rows: list[dict[str, Any]] = []
    for dataset, schema in CANONICAL_SCHEMAS.items():
        cfg = (draft.get("datasets") or {}).get(dataset) or {}
        source = cfg.get("source") or {}
        suggestions = cfg.get("_suggestions") or {}
        fields = list(dict.fromkeys(schema.get("required", []) + schema.get("recommended", [])))
        for field in fields:
            rule = (cfg.get("columns") or {}).get(field) or {}
            if isinstance(rule, str):
                rule = {"source": rule}
            suggestion = suggestions.get(field) or {}
            rows.append({
                "dataset": dataset,
                "enabled": bool(cfg.get("enabled", False)),
                "source_file": source.get("file"),
                "source_sheet": source.get("sheet"),
                "header_row": source.get("header", 0),
                "target_field": field,
                "required": field in schema.get("required", []),
                "source_field": rule.get("source") or suggestion.get("suggested_source"),
                "coalesce_fields": ";".join(rule.get("sources") or []),
                "concat_fields": ";".join(rule.get("concat") or []),
                "separator": rule.get("separator", " "),
                "date_source": rule.get("date_source"),
                "time_source": rule.get("time_source"),
                "constant": rule.get("constant"),
                "data_type": rule.get("type"),
                "default": rule.get("default"),
                "keep_unmapped": rule.get("keep_unmapped", True),
                "confidence": suggestion.get("confidence"),
                "notes": "",
            })
    return pd.DataFrame(rows)


def draft_to_file_context_table(draft: dict[str, Any]) -> pd.DataFrame:
    rows: list[dict[str, Any]] = []
    for item in _inventory(draft):
        role, confidence = _detected_role_for_item(draft, item["file"], item["sheet"])
        rows.append({
            "source_file": item["file"],
            "source_sheet": item["sheet"],
            "detected_role": role,
            "detected_confidence": confidence,
            "selected_role": role or "CONTEXT_ONLY",
            "use_as_primary_source": bool(role),
            "evidence_purpose": "",
            "period_start": "",
            "period_end": "",
            "document_status": "",
            "notes": "",
        })
    return pd.DataFrame(rows)


def draft_to_preview_table(draft: dict[str, Any]) -> pd.DataFrame:
    rows: list[dict[str, Any]] = []
    for dataset, schema in CANONICAL_SCHEMAS.items():
        cfg = (draft.get("datasets") or {}).get(dataset) or {}
        source = cfg.get("source") or {}
        mapped = cfg.get("columns") or {}
        required = list(schema.get("required", []))
        missing = [field for field in required if field not in mapped]
        rows.append({
            "dataset": dataset,
            "detected": bool(cfg.get("enabled", False)),
            "source_file": source.get("file"),
            "source_sheet": source.get("sheet"),
            "confidence": cfg.get("_dataset_confidence"),
            "mapped_fields": len(mapped),
            "required_fields": len(required),
            "missing_required_fields": ", ".join(missing),
            "preview_status": "READY_TO_REVIEW" if cfg.get("enabled") and not missing else ("MAPPING_INCOMPLETE" if cfg.get("enabled") else "NOT_DETECTED"),
        })
    return pd.DataFrame(rows)


def _excel_lists(draft: dict[str, Any]) -> dict[str, list[str]]:
    inventory = _inventory(draft)
    files = sorted({str(x["file"]) for x in inventory if x.get("file")})
    sheets = sorted({str(x["sheet"]) for x in inventory if x.get("sheet")})
    columns = sorted({str(col) for x in inventory for col in (x.get("columns") or []) if str(col).strip()})
    if not columns:
        fields = draft_to_field_table(draft)
        for col in ("source_field", "date_source", "time_source"):
            if col in fields.columns:
                columns.extend(fields[col].dropna().astype(str).tolist())
        columns = sorted(set(columns))
    return {
        "datasets": list(CANONICAL_SCHEMAS),
        "roles": list(CANONICAL_SCHEMAS) + SPECIAL_FILE_ROLES,
        "files": files,
        "sheets": sheets,
        "columns": columns,
        "types": ["string", "number", "integer", "date", "datetime", "boolean"],
        "bools": ["TRUE", "FALSE"],
    }


def _add_range_validation(workbook, worksheet, cell_range: str, values: list[str], list_name: str, *, allow_blank: bool = True) -> None:
    from openpyxl.worksheet.datavalidation import DataValidation

    if not values:
        return
    lists = workbook["_lists"]
    start_col = lists.max_column + 1 if lists.max_column else 1
    lists.cell(1, start_col, list_name)
    for idx, value in enumerate(values, start=2):
        lists.cell(idx, start_col, value)
    letter = lists.cell(1, start_col).column_letter
    formula = f"'_lists'!${letter}$2:${letter}${len(values) + 1}"
    validation = DataValidation(type="list", formula1=formula, allow_blank=allow_blank)
    worksheet.add_data_validation(validation)
    validation.add(cell_range)


def write_mapping_workbook(draft: dict[str, Any], path: str | Path) -> Path:
    """Write the controlled Advanced Mapping correction/context workbook.

    The workbook is optional: normal CSV/XLSX intake can proceed automatically.
    Operators use it only to correct AuditHero's interpretation or add context.
    Selection cells use controlled dropdown lists wherever practical so internal
    schema names and source headings do not have to be retyped.
    """
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    fields = draft_to_field_table(draft)
    file_context = draft_to_file_context_table(draft)
    preview = draft_to_preview_table(draft)
    values = pd.DataFrame(columns=["dataset", "target_field", "source_value", "target_value", "notes"])
    instructions = pd.DataFrame([
        ["AuditHero Advanced Mapping / Context Workbook"],
        ["Purpose", "Optional correction layer. The uploaded source files remain the audit evidence; this workbook only tells AuditHero how to interpret them."],
        ["Normal workflow", "Upload CSV/XLSX files, review AuditHero's preview, and continue when the interpretation is correct."],
        ["When to use this workbook", "Use it when a file role or field was missed, a source column was mapped incorrectly, or extra context is needed for reconciliation/storytelling."],
        ["file_context", "Confirm or correct what each uploaded file/sheet represents. Example: payroll_earnings means actual payroll earning lines used to confirm what was paid."],
        ["field_mapping", "Correct source-to-AuditHero field mappings. Use dropdowns rather than typing source headings where available."],
        ["value_mapping", "Translate source values into controlled AuditHero values when needed."],
        ["preview", "Shows the automatic interpretation that existed when this correction workbook was generated."],
        ["Safety", "No formulas or Python expressions are executed from this workbook. Original source files are never changed. Missing required mapped fields fail closed during conversion."],
        ["Not mandatory", "If automatic intake is correct and sufficient evidence is present, no Advanced Mapping workbook is required to run the audit."],
    ])
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        instructions.to_excel(writer, sheet_name="README", header=False, index=False)
        preview.to_excel(writer, sheet_name="preview", index=False)
        file_context.to_excel(writer, sheet_name="file_context", index=False)
        fields.to_excel(writer, sheet_name="field_mapping", index=False)
        values.to_excel(writer, sheet_name="value_mapping", index=False)

    from openpyxl import load_workbook

    wb = load_workbook(path)
    lists_ws = wb.create_sheet("_lists")
    lists_ws.sheet_state = "hidden"
    lists = _excel_lists(draft)

    ws = wb["field_mapping"]
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    widths = {
        "A": 28, "B": 10, "C": 34, "D": 26, "E": 12, "F": 34, "G": 10,
        "H": 34, "I": 34, "J": 34, "K": 12, "L": 30, "M": 30, "N": 22,
        "O": 14, "P": 18, "Q": 16, "R": 12, "S": 42,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    max_field_row = max(2, ws.max_row)
    _add_range_validation(wb, ws, f"A2:A{max_field_row}", lists["datasets"], "datasets", allow_blank=False)
    _add_range_validation(wb, ws, f"B2:B{max_field_row}", lists["bools"], "bools")
    _add_range_validation(wb, ws, f"C2:C{max_field_row}", lists["files"], "files")
    _add_range_validation(wb, ws, f"D2:D{max_field_row}", lists["sheets"], "sheets")
    for col in ("H", "L", "M"):
        _add_range_validation(wb, ws, f"{col}2:{col}{max_field_row}", lists["columns"], f"source_columns_{col}")
    _add_range_validation(wb, ws, f"O2:O{max_field_row}", lists["types"], "types")
    _add_range_validation(wb, ws, f"Q2:Q{max_field_row}", lists["bools"], "keep_unmapped")

    context = wb["file_context"]
    context.freeze_panes = "A2"
    context.auto_filter.ref = context.dimensions
    for col, width in {"A": 34, "B": 26, "C": 28, "D": 20, "E": 30, "F": 22, "G": 48, "H": 16, "I": 16, "J": 24, "K": 48}.items():
        context.column_dimensions[col].width = width
    max_context_row = max(2, context.max_row)
    _add_range_validation(wb, context, f"E2:E{max_context_row}", lists["roles"], "file_roles")
    _add_range_validation(wb, context, f"F2:F{max_context_row}", lists["bools"], "context_bools")

    preview_ws = wb["preview"]
    preview_ws.freeze_panes = "A2"
    preview_ws.auto_filter.ref = preview_ws.dimensions
    for col, width in {"A": 30, "B": 12, "C": 34, "D": 24, "E": 12, "F": 14, "G": 16, "H": 54, "I": 24}.items():
        preview_ws.column_dimensions[col].width = width

    vm = wb["value_mapping"]
    vm.freeze_panes = "A2"
    vm.auto_filter.ref = vm.dimensions
    for col, width in {"A": 28, "B": 34, "C": 30, "D": 30, "E": 40}.items():
        vm.column_dimensions[col].width = width
    _add_range_validation(wb, vm, "A2:A1000", lists["datasets"], "value_datasets")

    readme = wb["README"]
    readme.column_dimensions["A"].width = 30
    readme.column_dimensions["B"].width = 118
    wb.save(path)
    return path


def load_mapping_workbook(path: str | Path) -> dict[str, Any]:
    """Convert an edited Advanced Mapping workbook into the version-1 mapping dictionary."""
    path = Path(path)
    fields = pd.read_excel(path, sheet_name="field_mapping")
    try:
        values = pd.read_excel(path, sheet_name="value_mapping")
    except ValueError:
        values = pd.DataFrame(columns=["dataset", "target_field", "source_value", "target_value"])
    try:
        file_context = pd.read_excel(path, sheet_name="file_context")
    except ValueError:
        file_context = pd.DataFrame()

    primary_sources: dict[str, dict[str, Any]] = {}
    if not file_context.empty:
        for _, row in file_context.iterrows():
            role = _text(row.get("selected_role"))
            if role not in CANONICAL_SCHEMAS or not _bool(row.get("use_as_primary_source"), False):
                continue
            primary_sources[role] = {
                "file": _text(row.get("source_file")),
                "sheet": _text(row.get("source_sheet")),
                "header": 0,
            }

    datasets: dict[str, Any] = {}
    for dataset, group in fields.groupby("dataset", dropna=True):
        dataset = str(dataset).strip()
        if dataset not in CANONICAL_SCHEMAS:
            continue
        first = group.iloc[0]
        source = primary_sources.get(dataset) or {
            "file": _text(first.get("source_file")),
            "sheet": _text(first.get("source_sheet")),
            "header": int(first.get("header_row") or 0),
        }
        cfg: dict[str, Any] = {
            "enabled": _bool(first.get("enabled"), False),
            "source": source,
            "columns": {},
        }
        if dataset in primary_sources:
            cfg["enabled"] = True
        for _, row in group.iterrows():
            target = _text(row.get("target_field"))
            if not target or target not in set(CANONICAL_SCHEMAS[dataset].get("required", []) + CANONICAL_SCHEMAS[dataset].get("recommended", [])):
                continue
            rule: dict[str, Any] = {}
            if _text(row.get("source_field")):
                rule["source"] = _text(row.get("source_field"))
            elif _split(row.get("coalesce_fields")):
                rule["sources"] = _split(row.get("coalesce_fields"))
            elif _split(row.get("concat_fields")):
                rule["concat"] = _split(row.get("concat_fields"))
                rule["separator"] = _text(row.get("separator")) or " "
            elif _text(row.get("date_source")) and _text(row.get("time_source")):
                rule["date_source"] = _text(row.get("date_source"))
                rule["time_source"] = _text(row.get("time_source"))
            elif _text(row.get("constant")) is not None:
                rule["constant"] = row.get("constant")
            else:
                continue
            dtype = _text(row.get("data_type"))
            if dtype:
                rule["type"] = dtype
            default = row.get("default")
            if default is not None and not (isinstance(default, float) and pd.isna(default)):
                rule["default"] = default
            rule["keep_unmapped"] = _bool(row.get("keep_unmapped"), True)

            if not values.empty:
                subset = values[
                    (values["dataset"].astype(str).str.strip() == dataset)
                    & (values["target_field"].astype(str).str.strip() == target)
                ]
                value_map = {}
                for _, vrow in subset.iterrows():
                    source_value = vrow.get("source_value")
                    target_value = vrow.get("target_value")
                    if pd.notna(source_value):
                        value_map[source_value] = target_value
                if value_map:
                    rule["value_map"] = value_map
            cfg["columns"][target] = rule
        datasets[dataset] = cfg

    context_records: list[dict[str, Any]] = []
    if not file_context.empty:
        for _, row in file_context.iterrows():
            context_records.append({
                "source_file": _text(row.get("source_file")),
                "source_sheet": _text(row.get("source_sheet")),
                "selected_role": _text(row.get("selected_role")),
                "evidence_purpose": _text(row.get("evidence_purpose")),
                "period_start": _text(row.get("period_start")),
                "period_end": _text(row.get("period_end")),
                "document_status": _text(row.get("document_status")),
                "notes": _text(row.get("notes")),
            })
    return {"version": 1, "datasets": datasets, "file_context": context_records, "source": str(path)}


def load_mapping(path: str | Path) -> dict[str, Any]:
    path = Path(path)
    if path.suffix.lower() == ".json":
        return json.loads(path.read_text(encoding="utf-8"))
    if path.suffix.lower() in {".xlsx", ".xlsm"}:
        return load_mapping_workbook(path)
    raise ValueError("Mapping must be a .json, .xlsx or .xlsm file")
