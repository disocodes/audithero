from pathlib import Path

import pandas as pd
import pytest
from openpyxl import load_workbook

from schads_audit.mapping_workbook import write_mapping_workbook, load_mapping_workbook


def _draft():
    return {
        "version": 1,
        "_source_inventory": [
            {"file": "employees.csv", "sheet": None, "item_name": "employees", "sample_rows": 2, "columns": ["Employee Number", "Full Name", "Employment Type"]},
            {"file": "payroll.xlsx", "sheet": "Payroll Detail", "item_name": "payroll Payroll Detail", "sample_rows": 5, "columns": ["Employee Number", "Period Start", "Period End", "Pay Item", "Amount"]},
            {"file": "payroll_adjustment.xlsx", "sheet": "Adjustment", "item_name": "payroll adjustment", "sample_rows": 2, "columns": ["Employee Number", "Period Start", "Period End", "Pay Item", "Amount"]},
        ],
        "datasets": {
            "employees": {
                "enabled": True,
                "source": {"file": "employees.csv", "sheet": None, "header": 0},
                "columns": {"employee_id": {"source": "Employee Number"}, "employee_name": {"source": "Full Name"}},
                "_suggestions": {"employee_id": {"suggested_source": "Employee Number", "confidence": 1.0}, "employee_name": {"suggested_source": "Full Name", "confidence": 1.0}},
                "_dataset_confidence": 0.99,
            },
            "payroll_earnings": {
                "enabled": True,
                "source": {"file": "payroll.xlsx", "sheet": "Payroll Detail", "header": 0},
                "columns": {
                    "employee_id": {"source": "Employee Number"},
                    "pay_period_start": {"source": "Period Start"},
                    "pay_period_end": {"source": "Period End"},
                    "pay_category": {"source": "Pay Item"},
                    "amount": {"source": "Amount"},
                },
                "_suggestions": {},
                "_dataset_confidence": 0.95,
            },
        },
    }


def test_mapping_workbook_round_trip(tmp_path: Path):
    path = tmp_path / "source_mapping.xlsx"
    write_mapping_workbook(_draft(), path)
    assert path.exists()

    fields = pd.read_excel(path, sheet_name="field_mapping")
    values = pd.DataFrame([{"dataset": "employees", "target_field": "employment_type_current", "source_value": "Permanent FT", "target_value": "FULL_TIME", "notes": ""}])
    readme = pd.read_excel(path, sheet_name="README", header=None)
    file_context = pd.read_excel(path, sheet_name="file_context")
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        readme.to_excel(writer, sheet_name="README", header=False, index=False)
        fields.to_excel(writer, sheet_name="field_mapping", index=False)
        values.to_excel(writer, sheet_name="value_mapping", index=False)
        file_context.to_excel(writer, sheet_name="file_context", index=False)

    mapping = load_mapping_workbook(path)
    assert mapping["version"] == 1
    assert mapping["datasets"]["employees"]["enabled"] is True
    assert mapping["datasets"]["employees"]["columns"]["employee_id"]["source"] == "Employee Number"
    assert mapping["file_context"]


def test_advanced_mapping_workbook_has_preview_context_and_dropdowns(tmp_path: Path):
    path = tmp_path / "source_mapping_draft.xlsx"
    write_mapping_workbook(_draft(), path)
    wb = load_workbook(path)

    assert {"README", "preview", "file_context", "field_mapping", "value_mapping", "_lists"}.issubset(wb.sheetnames)
    assert wb["_lists"].sheet_state == "hidden"
    assert "file_roles" in wb.defined_names
    assert "source_columns_H" in wb.defined_names

    preview = pd.read_excel(path, sheet_name="preview")
    payroll_preview = preview[preview["dataset"] == "payroll_earnings"].iloc[0]
    assert bool(payroll_preview["detected"]) is True
    assert payroll_preview["source_file"] == "payroll.xlsx"
    assert payroll_preview["preview_status"] == "READY_TO_REVIEW"

    context = pd.read_excel(path, sheet_name="file_context")
    payroll_context = context[context["source_file"] == "payroll.xlsx"].iloc[0]
    assert payroll_context["detected_role"] == "payroll_earnings"
    assert payroll_context["selected_role"] == "payroll_earnings"

    context_validations = list(wb["file_context"].data_validations.dataValidation)
    field_validations = list(wb["field_mapping"].data_validations.dataValidation)
    assert any("E2" in str(v.sqref) for v in context_validations)
    assert any("H2" in str(v.sqref) for v in field_validations)


def test_file_context_can_correct_primary_payroll_source(tmp_path: Path):
    path = tmp_path / "source_mapping.xlsx"
    write_mapping_workbook(_draft(), path)
    wb = load_workbook(path)
    ws = wb["file_context"]
    headers = {cell.value: cell.column for cell in ws[1]}
    payroll_row = next(row for row in range(2, ws.max_row + 1) if ws.cell(row, headers["source_file"]).value == "payroll.xlsx")
    ws.cell(payroll_row, headers["selected_role"], "payroll_earnings")
    ws.cell(payroll_row, headers["use_as_primary_source"], "TRUE")
    wb.save(path)

    mapping = load_mapping_workbook(path)
    payroll = mapping["datasets"]["payroll_earnings"]
    assert payroll["enabled"] is True
    assert payroll["source"]["file"] == "payroll.xlsx"
    assert payroll["source"]["sheet"] == "Payroll Detail"


def test_context_only_disables_original_detected_source(tmp_path: Path):
    path = tmp_path / "source_mapping.xlsx"
    write_mapping_workbook(_draft(), path)
    wb = load_workbook(path)
    ws = wb["file_context"]
    headers = {cell.value: cell.column for cell in ws[1]}
    row = next(r for r in range(2, ws.max_row + 1) if ws.cell(r, headers["source_file"]).value == "payroll.xlsx")
    ws.cell(row, headers["selected_role"], "CONTEXT_ONLY")
    ws.cell(row, headers["use_as_primary_source"], "FALSE")
    wb.save(path)

    mapping = load_mapping_workbook(path)
    assert mapping["datasets"]["payroll_earnings"]["enabled"] is False


def test_duplicate_primary_sources_are_rejected(tmp_path: Path):
    path = tmp_path / "source_mapping.xlsx"
    write_mapping_workbook(_draft(), path)
    wb = load_workbook(path)
    ws = wb["file_context"]
    headers = {cell.value: cell.column for cell in ws[1]}
    for filename in ("payroll.xlsx", "payroll_adjustment.xlsx"):
        row = next(r for r in range(2, ws.max_row + 1) if ws.cell(r, headers["source_file"]).value == filename)
        ws.cell(row, headers["selected_role"], "payroll_earnings")
        ws.cell(row, headers["use_as_primary_source"], "TRUE")
    wb.save(path)

    with pytest.raises(ValueError, match="Multiple primary sources"):
        load_mapping_workbook(path)
