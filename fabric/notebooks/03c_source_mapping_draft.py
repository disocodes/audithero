# AuditHero Fabric notebook source
#
# PURPOSE
# -------
# Build an editable Excel mapping workbook from CSV/XLSX payroll, HR, roster and
# timekeeping exports. The notebook proposes matches to AuditHero's canonical
# input model and, when payroll earnings are supplied, lists the payroll earning
# categories that require an approved audit treatment. A pay category is a
# payroll earning or payroll item type such as Ordinary Hours, Overtime,
# Allowance or Leave. The notebook does not calculate payroll.

from pathlib import Path
import pandas as pd

try:
    from schads_audit.source_mapping import scan_source_items, generate_mapping_draft
    from schads_audit.mapping_workbook import write_mapping_workbook
except ModuleNotFoundError as exc:
    if exc.name == "schads_audit" or str(exc.name or "").startswith("schads_audit."):
        raise RuntimeError(
            "AuditHero is not available in the current Fabric Spark session. "
            "Confirm that AuditHero_Environment is attached to this notebook, then "
            "start a new Spark session and run the notebook again."
        ) from exc
    raise


def _normalize_fabric_lakehouse_path(value: str) -> str:
    """Normalize Fabric Lakehouse paths to the default notebook mount."""
    text = str(value or "").strip()
    for source, target in (
        ("/lakehouse/Files", "/lakehouse/default/Files"),
        ("/lakehouse/Tables", "/lakehouse/default/Tables"),
    ):
        if text == source or text.startswith(source + "/"):
            normalized = target + text[len(source):]
            print(f"Normalized Fabric Lakehouse path: {text} -> {normalized}")
            return normalized
    return text


def _suggest_pay_treatment(value: str) -> str:
    """Provide non-binding guidance for common pay-category labels."""
    text = str(value or "").strip().lower()
    if not text:
        return ""
    if any(term in text for term in ("annual leave", "personal leave", "sick leave", "long service leave", "leave without pay")):
        return "EXCLUDE"
    if any(term in text for term in ("allowance", "sleepover", "on call", "on-call", "meal allowance", "vehicle allowance")):
        return "ALLOWANCE"
    if any(term in text for term in ("ordinary", "overtime", "saturday", "sunday", "public holiday", "weekend", "penalty", "shift", "night", "afternoon")):
        return "AUDITABLE_WORK"
    return ""


def _mapped_source_field(cfg: dict, target: str) -> str | None:
    rule = (cfg.get("columns") or {}).get(target) or {}
    if isinstance(rule, str):
        return rule
    return rule.get("source")


def _validate_pay_category_source(cfg: dict, source_field: str, frame: pd.DataFrame) -> None:
    """Reject source fields that are inconsistent with payroll earning categories."""
    employee_field = _mapped_source_field(cfg, "employee_id")
    if employee_field and str(employee_field).strip() == str(source_field).strip():
        raise ValueError(
            "The payroll pay_category field has been mapped to the same source column as employee_id. "
            "Map pay_category to the payroll earning, item or category column."
        )

    field_name = str(source_field).strip().lower()
    employee_words = ("employee", "staff", "worker", "person")
    payroll_words = ("pay", "earning", "category", "item", "allowance", "hours", "rate")
    if any(word in field_name for word in employee_words) and not any(word in field_name for word in payroll_words):
        raise ValueError(
            f"The proposed pay_category source column '{source_field}' does not appear to be a payroll earning category field. "
            "Map pay_category to the payroll earning, item or category column."
        )

    if employee_field and employee_field in frame.columns and source_field in frame.columns:
        categories = set(frame[source_field].dropna().astype(str).str.strip())
        employees = set(frame[employee_field].dropna().astype(str).str.strip())
        categories.discard("")
        employees.discard("")
        if len(categories) >= 3 and len(employees) >= 3:
            overlap = len(categories & employees) / max(1, len(categories))
            if overlap >= 0.8:
                raise ValueError(
                    "The proposed pay_category values overlap substantially with employee identifiers. "
                    "Map pay_category to the payroll earning, item or category field."
                )


def _read_pay_categories(mapping: dict, raw_root: str) -> list[str]:
    """Read unique pay-category values from the detected payroll-earnings source."""
    cfg = (mapping.get("datasets") or {}).get("payroll_earnings") or {}
    if not cfg.get("enabled"):
        return []
    source = cfg.get("source") or {}
    source_field = _mapped_source_field(cfg, "pay_category")
    employee_field = _mapped_source_field(cfg, "employee_id")
    source_file = source.get("file")
    if not source_file or not source_field:
        return []

    source_path = Path(raw_root) / str(source_file)
    if not source_path.exists():
        return []
    usecols = [source_field]
    if employee_field and employee_field != source_field:
        usecols.append(employee_field)
    try:
        if source_path.suffix.lower() == ".csv":
            frame = pd.read_csv(source_path, usecols=lambda c: c in set(usecols))
        elif source_path.suffix.lower() in {".xlsx", ".xlsm"}:
            frame = pd.read_excel(source_path, sheet_name=source.get("sheet", 0), usecols=lambda c: c in set(usecols))
        else:
            return []
    except Exception as exc:
        print(f"Pay-category values could not be sampled automatically: {exc}")
        return []

    if source_field not in frame.columns:
        return []
    _validate_pay_category_source(cfg, source_field, frame)
    values = frame[source_field].dropna().astype(str).str.strip()
    return sorted(v for v in values.unique().tolist() if v)


def _add_pay_category_sheet(workbook_path: Path, pay_categories: list[str]) -> None:
    """Add the controlled pay-category treatment table to the mapping workbook."""
    from openpyxl import load_workbook
    from openpyxl.worksheet.datavalidation import DataValidation

    wb = load_workbook(workbook_path)
    if "pay_category_treatment" in wb.sheetnames:
        del wb["pay_category_treatment"]
    ws = wb.create_sheet("pay_category_treatment")
    ws.append(["pay_category", "treatment", "suggested_treatment", "definition", "notes"])
    for value in pay_categories:
        ws.append([
            value,
            "",
            _suggest_pay_treatment(value),
            "Payroll earning or payroll item type used for actual-pay reconciliation.",
            "Review and approve the treatment before conversion.",
        ])

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 24
    ws.column_dimensions["D"].width = 72
    ws.column_dimensions["E"].width = 56
    validation = DataValidation(
        type="list",
        formula1='"AUDITABLE_WORK,ALLOWANCE,EXCLUDE"',
        allow_blank=True,
    )
    ws.add_data_validation(validation)
    validation.add(f"B2:B{max(2, ws.max_row)}")

    readme = wb["README"]
    guidance = [
        ("pay_category_treatment", "Classifies each payroll earning or payroll item type for actual-pay reconciliation."),
        ("AUDITABLE_WORK", "Use for pay for worked hours, penalties or overtime that should count toward actual worked pay."),
        ("ALLOWANCE", "Use for a separate allowance payment that should count in the relevant entitlement comparison."),
        ("EXCLUDE", "Use for payroll items that should not count in the worked-pay comparison, such as leave or reimbursements."),
    ]
    for label, explanation in guidance:
        next_row = readme.max_row + 1
        readme.cell(next_row, 1, label)
        readme.cell(next_row, 2, explanation)
    wb.save(workbook_path)


# Parameters supplied by the Fabric pipeline. They can also be set when the
# notebook is run directly.
source_root = _normalize_fabric_lakehouse_path(source_root)
draft_path = _normalize_fabric_lakehouse_path(draft_path)

print("STEP 1 — Inspect raw source files")
source_dir = Path(source_root)
if not source_dir.exists():
    source_dir.mkdir(parents=True, exist_ok=True)
    print(f"Created AuditHero raw-import folder: {source_root}")

inventory = scan_source_items(source_root)
if inventory.empty:
    raise ValueError(
        f"No CSV/XLSX files were found under {source_root}. "
        "Upload the payroll, HR, roster or timekeeping exports to this Lakehouse Files folder "
        "and run this notebook again."
    )
display(inventory[["file", "sheet", "item_name", "sample_rows", "columns"]])

print("STEP 2 — Suggest canonical dataset and field matches")
draft_file = Path(draft_path)
draft_file.parent.mkdir(parents=True, exist_ok=True)
if draft_file.exists() and str(overwrite).lower() not in {"true", "1", "yes"}:
    raise FileExistsError(
        f"{draft_path} already exists. Set overwrite=true only when the existing draft should be replaced."
    )

draft = generate_mapping_draft(source_root)
if "pay_category_mapping" in (draft.get("datasets") or {}):
    draft["datasets"]["pay_category_mapping"]["enabled"] = False
    draft["datasets"]["pay_category_mapping"]["source"] = None
    draft["datasets"]["pay_category_mapping"]["columns"] = {}

pay_categories = _read_pay_categories(draft, source_root)
write_mapping_workbook(draft, draft_file)
_add_pay_category_sheet(draft_file, pay_categories)

summary_rows = []
for dataset, cfg in draft["datasets"].items():
    source = cfg.get("source") or {}
    summary_rows.append({
        "dataset": dataset,
        "suggested": bool(cfg.get("enabled")),
        "source_file": source.get("file"),
        "source_sheet": source.get("sheet"),
        "confidence": cfg.get("_dataset_confidence"),
        "suggested_fields": len(cfg.get("columns") or {}),
    })
summary = pd.DataFrame(summary_rows)
display(summary.sort_values(["suggested", "confidence"], ascending=[False, False]))

if pay_categories:
    display(pd.DataFrame({
        "pay_category": pay_categories,
        "suggested_treatment": [_suggest_pay_treatment(value) for value in pay_categories],
    }))
    print(f"Detected {len(pay_categories)} unique payroll earning category value(s). Complete the pay_category_treatment sheet before conversion.")

print("STEP 3 — Review and approve the mapping")
print(f"Mapping draft created: {draft_path}")
print("Download the workbook from the Lakehouse Files area and review field_mapping and value_mapping as required.")
if pay_categories:
    print("Also review pay_category_treatment. Each row represents a payroll earning or payroll item type found in the source payroll data.")
print("Upload the approved workbook as source_mapping.xlsx.")
print("NEXT: run 'AuditHero - Convert Mapped Files and Run Audit'.")
print("Use 'AuditHero - Convert Source Files' when conversion and File Readiness need to be checked without running the payroll audit.")

notebookutils.notebook.exit(draft_path)
