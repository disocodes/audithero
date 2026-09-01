# Databricks notebook source
# MAGIC %md
# MAGIC # AuditHero — Build a Source Mapping Workbook
# MAGIC
# MAGIC **Purpose:** create an editable source-mapping workbook when payroll, HR, roster or timekeeping exports do not already use AuditHero's canonical field names.
# MAGIC
# MAGIC The notebook scans CSV and Excel files in the raw import folder, compares file/sheet names and column headings with AuditHero's canonical fields, and creates `source_mapping_draft.xlsx`. When usable payroll earning-line detail is detected, it also lists the payroll earning categories that require an approved treatment before actual-pay reconciliation can run.
# MAGIC
# MAGIC A **pay category** is a payroll earning or payroll item type such as Ordinary Hours, Saturday, Overtime, Sleepover Allowance or Annual Leave.
# MAGIC
# MAGIC **Workflow:** upload raw exports → run this notebook → review and save `source_mapping.xlsx` → run **AuditHero - Convert Mapped Files and Run Audit**.
# COMMAND ----------
# MAGIC %pip install "openpyxl>=3.1"
# COMMAND ----------
# MAGIC %md
# MAGIC ## 1. Read Job parameters
# MAGIC
# MAGIC The parameters are exposed by the Databricks Job and can also be set when this notebook is run directly. Blank path values use the standard AuditHero Unity Catalog Volume locations.
# COMMAND ----------
from pathlib import Path
import shutil
import sys
import tempfile

import pandas as pd

ROOT = Path.cwd()
for candidate in (ROOT, *ROOT.parents):
    if (candidate / "src" / "schads_audit").exists():
        sys.path.insert(0, str(candidate / "src"))
        break

from schads_audit.source_mapping import scan_source_items, generate_mapping_draft
from schads_audit.source_mapping_hardening import harden_payroll_earnings_draft
from schads_audit.mapping_workbook import write_mapping_workbook

# Job parameters. `catalog` determines the default Volume paths.
dbutils.widgets.text("catalog", "schads_payroll")
dbutils.widgets.text("source_root", "")
dbutils.widgets.text("draft_path", "")
dbutils.widgets.dropdown("overwrite", "false", ["false", "true"])

catalog = dbutils.widgets.get("catalog")
source_root = dbutils.widgets.get("source_root").strip() or f"/Volumes/{catalog}/bronze/landing/import/raw"
draft_path = dbutils.widgets.get("draft_path").strip() or f"/Volumes/{catalog}/bronze/landing/import/source_mapping_draft.xlsx"
overwrite = dbutils.widgets.get("overwrite").lower() == "true"

print(f"Raw source folder: {source_root}")
print(f"Mapping draft:     {draft_path}")
# COMMAND ----------
# MAGIC %md
# MAGIC ## 2. Inspect uploaded files and sheets
# MAGIC
# MAGIC `scan_source_items` reads a small sample of each CSV file or Excel sheet to identify the source structure and column headings.
# COMMAND ----------
inventory = scan_source_items(source_root)
if inventory.empty:
    raise ValueError(
        f"No CSV/XLSX files were found under {source_root}. Upload the source exports in Catalog Explorer, then run this job again."
    )
display(inventory[["file", "sheet", "item_name", "sample_rows", "columns"]])
# COMMAND ----------
# MAGIC %md
# MAGIC ## 3. Generate mapping suggestions
# MAGIC
# MAGIC AuditHero compares source headings with the canonical schema and proposes dataset and field matches. Review every required field before conversion, especially classifications, employment type, dates and payroll amounts.
# MAGIC
# MAGIC Payroll earning-line detection uses stricter header rules than the general mapper. Actual-pay reconciliation is enabled only when a source exposes a trustworthy employee identifier, pay-period start/end, payroll earning category and amount. If those fields are not available, the entitlement audit can still run without actual-pay reconciliation.
# MAGIC
# MAGIC If payroll earnings are available, review the `pay_category_treatment` sheet as well. Each row represents a payroll earning or payroll item type found in the source payroll data.
# MAGIC
# MAGIC Treatment meanings:
# MAGIC - `AUDITABLE_WORK` — pay for worked hours, penalties or overtime that should count toward actual worked pay.
# MAGIC - `ALLOWANCE` — a separate allowance payment that should count in the relevant entitlement comparison.
# MAGIC - `EXCLUDE` — a payroll item that should not be counted in the worked-pay comparison, such as leave or a reimbursement.
# MAGIC
# MAGIC Suggested treatments are guidance only; the `treatment` column is the approved control used by the audit.
# COMMAND ----------
def _volume_uri(path: str) -> str:
    return f"dbfs:{path}" if path.startswith("/Volumes/") else path


def _path_exists(path: str) -> bool:
    if path.startswith("/Volumes/"):
        try:
            dbutils.fs.ls(_volume_uri(path))
            return True
        except Exception:
            return False
    return Path(path).exists()


def _suggest_pay_treatment(value: str) -> str:
    """Provide non-binding guidance for common pay-category labels."""
    text = str(value or "").strip().lower()
    if not text:
        return ""
    if any(term in text for term in ("annual leave", "personal leave", "sick leave", "long service leave", "leave without pay")):
        return "EXCLUDE"
    if any(term in text for term in ("allowance", "sleepover", "on call", "on-call", "meal allowance", "vehicle allowance")):
        return "ALLOWANCE"
    if any(term in text for term in ("ordinary", "overtime", "saturday", "sunday", "public holiday", "weekend", "penalty", "shift", "night", "afternoon")):
        return "AUDITABLE_WORK"
    return ""


def _mapped_source_field(cfg: dict, target: str) -> str | None:
    rule = (cfg.get("columns") or {}).get(target) or {}
    if isinstance(rule, str):
        return rule
    return rule.get("source")


def _validate_pay_category_source(cfg: dict, source_field: str, frame: pd.DataFrame) -> None:
    """Reject source fields that are clearly inconsistent with payroll earning categories."""
    employee_field = _mapped_source_field(cfg, "employee_id")
    if employee_field and str(employee_field).strip() == str(source_field).strip():
        raise ValueError(
            "The payroll pay_category field has been mapped to the same source column as employee_id. "
            "Map pay_category to the payroll earning, item or category column."
        )

    field_name = str(source_field).strip().lower()
    employee_words = ("employee", "staff", "worker", "person")
    payroll_words = ("pay", "earning", "category", "item", "allowance", "hours", "rate")
    if any(word in field_name for word in employee_words) and not any(word in field_name for word in payroll_words):
        raise ValueError(
            f"The proposed pay_category source column '{source_field}' does not appear to be a payroll earning category field. "
            "Map pay_category to the payroll earning, item or category column."
        )


def _read_pay_categories(mapping: dict) -> list[str]:
    """Read unique source pay categories from the detected payroll-earnings source."""
    cfg = (mapping.get("datasets") or {}).get("payroll_earnings") or {}
    if not cfg.get("enabled"):
        return []
    source = cfg.get("source") or {}
    source_field = _mapped_source_field(cfg, "pay_category")
    employee_field = _mapped_source_field(cfg, "employee_id")
    source_file = source.get("file")
    if not source_file or not source_field:
        return []

    source_path = Path(source_root) / str(source_file)
    if not source_path.exists():
        return []
    usecols = [source_field]
    if employee_field and employee_field != source_field:
        usecols.append(employee_field)
    try:
        if source_path.suffix.lower() == ".csv":
            frame = pd.read_csv(source_path, usecols=lambda c: c in set(usecols))
        elif source_path.suffix.lower() in {".xlsx", ".xlsm"}:
            frame = pd.read_excel(source_path, sheet_name=source.get("sheet", 0), usecols=lambda c: c in set(usecols))
        else:
            return []
    except Exception as exc:
        print(f"Pay-category values could not be sampled automatically: {exc}")
        return []

    if source_field not in frame.columns:
        return []
    _validate_pay_category_source(cfg, source_field, frame)
    values = frame[source_field].dropna().astype(str).str.strip()
    return sorted(v for v in values.unique().tolist() if v)


def _add_pay_category_sheet(workbook_path: Path, pay_categories: list[str]) -> None:
    """Add the controlled pay-category treatment table to the mapping workbook."""
    from openpyxl import load_workbook
    from openpyxl.worksheet.datavalidation import DataValidation

    wb = load_workbook(workbook_path)
    if "pay_category_treatment" in wb.sheetnames:
        del wb["pay_category_treatment"]
    ws = wb.create_sheet("pay_category_treatment")
    ws.append(["pay_category", "treatment", "suggested_treatment", "definition", "notes"])
    for value in pay_categories:
        ws.append([
            value,
            "",
            _suggest_pay_treatment(value),
            "Payroll earning or payroll item type used for actual-pay reconciliation.",
            "Review and approve the treatment before conversion.",
        ])

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 24
    ws.column_dimensions["D"].width = 72
    ws.column_dimensions["E"].width = 56
    treatment_validation = DataValidation(
        type="list",
        formula1='"AUDITABLE_WORK,ALLOWANCE,EXCLUDE"',
        allow_blank=True,
    )
    ws.add_data_validation(treatment_validation)
    treatment_validation.add(f"B2:B{max(2, ws.max_row)}")

    readme = wb["README"]
    guidance = [
        ("pay_category_treatment", "Classifies each payroll earning or payroll item type for actual-pay reconciliation."),
        ("AUDITABLE_WORK", "Use for pay for worked hours, penalties or overtime that should count toward actual worked pay."),
        ("ALLOWANCE", "Use for a separate allowance payment that should count in the relevant entitlement comparison."),
        ("EXCLUDE", "Use for payroll items that should not count in the worked-pay comparison, such as leave or reimbursements."),
    ]
    for label, explanation in guidance:
        next_row = readme.max_row + 1
        readme.cell(next_row, 1, label)
        readme.cell(next_row, 2, explanation)
    wb.save(workbook_path)


def _write_mapping_draft(mapping: dict, destination: str, pay_categories: list[str]) -> None:
    """Create the Excel workbook locally, then publish it to the configured destination."""
    if not destination.startswith("/Volumes/"):
        target = Path(destination)
        target.parent.mkdir(parents=True, exist_ok=True)
        write_mapping_workbook(mapping, target)
        _add_pay_category_sheet(target, pay_categories)
        return

    temp_dir = Path(tempfile.mkdtemp(prefix="audithero-mapping-"))
    local_file = temp_dir / Path(destination).name
    try:
        write_mapping_workbook(mapping, local_file)
        _add_pay_category_sheet(local_file, pay_categories)
        destination_parent = destination.rsplit("/", 1)[0]
        dbutils.fs.mkdirs(_volume_uri(destination_parent))
        shutil.copyfile(local_file, destination)
        if not Path(destination).exists() or Path(destination).stat().st_size == 0:
            raise OSError(f"Databricks could not publish the mapping workbook to {destination}")
    finally:
        shutil.rmtree(temp_dir, ignore_errors=True)


if _path_exists(draft_path) and not overwrite:
    raise FileExistsError(
        f"{draft_path} already exists. Set overwrite=true if the existing draft should be replaced."
    )

draft = harden_payroll_earnings_draft(generate_mapping_draft(source_root), source_root)
if "pay_category_mapping" in (draft.get("datasets") or {}):
    draft["datasets"]["pay_category_mapping"]["enabled"] = False
    draft["datasets"]["pay_category_mapping"]["source"] = None
    draft["datasets"]["pay_category_mapping"]["columns"] = {}
pay_categories = _read_pay_categories(draft)
_write_mapping_draft(draft, draft_path, pay_categories)

summary_rows = []
for dataset, cfg in draft["datasets"].items():
    source = cfg.get("source") or {}
    summary_rows.append({
        "dataset": dataset,
        "suggested": bool(cfg.get("enabled")),
        "source_file": source.get("file"),
        "source_sheet": source.get("sheet"),
        "confidence": cfg.get("_dataset_confidence"),
        "mapped_fields": len(cfg.get("columns") or {}),
    })

summary = pd.DataFrame(summary_rows)
display(summary.sort_values(["suggested", "confidence"], ascending=[False, False]))
if pay_categories:
    display(pd.DataFrame({
        "pay_category": pay_categories,
        "suggested_treatment": [_suggest_pay_treatment(value) for value in pay_categories],
    }))
    print(f"Detected {len(pay_categories)} unique payroll earning category value(s). Complete the pay_category_treatment sheet before conversion.")
else:
    payroll_cfg = (draft.get("datasets") or {}).get("payroll_earnings") or {}
    if payroll_cfg.get("_actual_pay_status") == "UNAVAILABLE":
        print("Actual-pay earning-line detail was not detected with sufficient confidence. Entitlement calculation can continue without actual-pay reconciliation.")
# COMMAND ----------
# MAGIC %md
# MAGIC ## 4. Review and approve the mapping
# MAGIC
# MAGIC In **Catalog Explorer**, download `source_mapping_draft.xlsx`, review the `field_mapping` sheet and add source-value translations on `value_mapping` where required.
# MAGIC
# MAGIC If payroll earnings are available, also review `pay_category_treatment`. Each row represents a payroll earning or payroll item type found in the source payroll data.
# MAGIC
# MAGIC Upload the approved workbook as `source_mapping.xlsx`. Missing optional actual-pay detail does not prevent entitlement calculation; missing required core audit evidence still stops conversion or readiness.
# COMMAND ----------
print("Mapping draft created successfully.")
print(draft_path)
print("NEXT: review the workbook, upload it as source_mapping.xlsx, then run 'AuditHero - Convert Mapped Files and Run Audit'.")
print("Use 'AuditHero - Convert Source Files' when conversion and File Readiness need to be checked without running the payroll audit.")
